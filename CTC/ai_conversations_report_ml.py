#!/usr/bin/env python3
"""
AI Conversations + Excel Report Machine Learning Engine
======================================================

Purpose
-------
Learns reporting structure from Excel workbooks and optional conversation history
(`ai_conversations.html`), maps report components/metrics, validates numerical
relationships, and generates coherent synthetic report scenarios at different
visual-similarity levels (90/60/30/10).

The engine deliberately separates:
1) ML discovery/classification (unsupervised text clustering + anomaly detection), and
2) financial/project-control arithmetic (deterministic constrained calculations).

This keeps generated reports realistic: CPI, CV, gross profit, cost mix, and cash flow
reconcile instead of being independently randomized.

Typical use
-----------
    python ai_conversations_report_ml.py analyze "Gloria Cost Report 06.2026.xlsx"
    python ai_conversations_report_ml.py learn "Gloria Cost Report 06.2026.xlsx" --html ai_conversations.html
    python ai_conversations_report_ml.py generate "Gloria Cost Report 06.2026.xlsx" --html ai_conversations.html --out generated_reports

Run without arguments for interactive mode.

Dependencies
------------
Required for ML: scikit-learn
Required for XLSX generation: openpyxl
Optional: joblib (model persistence), jinja2 (HTML audit output)

No internet access is required.
"""

from __future__ import annotations

import argparse
import csv
import difflib
import hashlib
import html
import json
import math
import os
import random
import re
import statistics
import sys
import tempfile
import textwrap
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict, field
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
import xml.etree.ElementTree as ET

APP_NAME = "AI Conversations Report ML Engine"
VERSION = "1.0.0"
DEFAULT_HTML = "ai_conversations.html"
MODEL_FILENAME = "report_ml_model.joblib"
MODEL_META_FILENAME = "report_ml_model.json"

NS_MAIN = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
NS_REL = {"r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
NS_PKG_REL = {"p": "http://schemas.openxmlformats.org/package/2006/relationships"}

SIMILARITY_PROFILES = {
    90: {
        "name": "Near Original",
        "description": "Retains the original reporting hierarchy with modest rearrangement.",
        "layout": "matrix",
        "chart_mix": ["column", "doughnut", "line", "bar"],
        "metric_density": 1.00,
    },
    60: {
        "name": "Executive Recomposition",
        "description": "Same reporting meaning, but reorganized into KPI cards and ranked drivers.",
        "layout": "executive",
        "chart_mix": ["bar", "line", "doughnut", "column"],
        "metric_density": 0.92,
    },
    30: {
        "name": "Control Room",
        "description": "Signals-first layout with exception ranking and component performance.",
        "layout": "control_room",
        "chart_mix": ["bar", "line", "scatter", "column"],
        "metric_density": 0.82,
    },
    10: {
        "name": "Decision Lens",
        "description": "A different visual family centered on management actions and exceptions.",
        "layout": "decision_lens",
        "chart_mix": ["line", "bar", "column"],
        "metric_density": 0.72,
    },
}

# Canonical project-control / cost-control vocabulary.
# Each concept includes synonyms to improve matching across different templates.
METRIC_LEXICON: Dict[str, List[str]] = {
    "contract_price": ["contract price", "contract value", "award value", "contract amount", "revised contract value"],
    "direct_budget": ["direct budget", "direct cost budget", "budget direct", "direct costs budget"],
    "indirect_budget": ["indirect budget", "indirect cost budget", "project overhead budget", "overheads budget"],
    "fees_budget": ["fees", "sponsorship", "fees and sponsorship", "commercial provisions", "deductions"],
    "total_budget": ["total budget", "project budget", "budget at completion", "bac", "total cost budget"],
    "earned_value": ["earned value", "ev", "budgeted cost of work performed", "bcwp"],
    "actual_cost": ["actual cost", "ac", "actual expense", "cost incurred", "acwp"],
    "cost_variance": ["cost variance", "cv", "variance cost"],
    "cpi": ["cpi", "cost performance index", "cost index"],
    "completion": ["completion", "completion %", "progress", "progress %", "physical progress"],
    "activity_weight": ["activity weight", "weight", "weight %", "wbs weight", "wp weight"],
    "revenue": ["revenue", "billed revenue", "unbilled revenue", "certified revenue", "sales"],
    "gross_profit": ["gross profit", "gp", "gross margin", "profit"],
    "gross_profit_pct": ["gross profit %", "gp %", "gross margin %", "margin %"],
    "cash_in": ["cash in", "cash inflow", "receipts", "collections"],
    "cash_out": ["cash out", "cash outflow", "payments", "disbursements"],
    "cash_flow": ["cash flow", "cumulative cash flow", "net cash"],
    "subcontractor_cost": ["subcontractor", "subcontract", "subcontractors cost"],
    "labor_cost": ["labor", "labour", "manpower cost", "labor cost"],
    "equipment_cost": ["equipment", "plant cost", "machinery cost"],
    "material_cost": ["material", "materials", "material cost"],
    "other_cost": ["other cost", "miscellaneous", "misc cost"],
    "soil_settlement": ["soil settlement", "settlement", "soil"],
    "wbs": ["wbs", "work breakdown structure", "control account", "work package"],
    "activity": ["activity", "activity id", "activity name", "task"],
    "cost_code": ["cost code", "cost element", "sap cost code", "gl account"],
    "status": ["status", "budget status", "signal", "performance status"],
}

# Arithmetic/semantic dependency graph used for validation and synthesis.
DEPENDENCY_GRAPH = {
    "total_budget": ["direct_budget", "indirect_budget", "fees_budget"],
    "cost_variance": ["earned_value", "actual_cost"],
    "cpi": ["earned_value", "actual_cost"],
    "gross_profit": ["revenue", "actual_cost"],
    "gross_profit_pct": ["gross_profit", "revenue"],
    "cash_flow": ["cash_in", "cash_out"],
}


@dataclass
class CellRecord:
    sheet: str
    ref: str
    row: int
    col: int
    value: Any = None
    formula: Optional[str] = None
    data_type: Optional[str] = None


@dataclass
class SheetProfile:
    name: str
    row_count: int = 0
    col_count: int = 0
    nonempty_cells: int = 0
    numeric_cells: int = 0
    formula_cells: int = 0
    text_cells: int = 0
    chart_count: int = 0
    drawing_count: int = 0
    numeric_density: float = 0.0
    formula_density: float = 0.0
    text_sample: str = ""
    discovered_metrics: List[str] = field(default_factory=list)
    cluster: Optional[int] = None
    anomaly_score: Optional[float] = None


@dataclass
class WorkbookProfile:
    path: str
    sha256: str
    sheet_names: List[str]
    sheets: List[SheetProfile]
    metrics: Dict[str, Dict[str, Any]]
    formulas: Dict[str, int]
    conversation_terms: List[str] = field(default_factory=list)
    learned_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: List[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]) -> None:
        if tag.lower() in {"script", "style", "svg", "noscript"}:
            self._skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "svg", "noscript"} and self._skip_depth:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if self._skip_depth == 0:
            s = re.sub(r"\s+", " ", html.unescape(data)).strip()
            if s:
                self.parts.append(s)

    def text(self) -> str:
        return "\n".join(self.parts)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).lower().strip()
    s = re.sub(r"[_\-–—/\\|]+", " ", s)
    s = re.sub(r"[^\w%+. ]+", " ", s, flags=re.UNICODE)
    return re.sub(r"\s+", " ", s).strip()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def column_index(ref: str) -> int:
    m = re.match(r"([A-Za-z]+)", ref)
    if not m:
        return 0
    n = 0
    for ch in m.group(1).upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def row_index(ref: str) -> int:
    m = re.search(r"(\d+)$", ref)
    return int(m.group(1)) if m else 0


def load_conversation_html(path: Optional[Path]) -> Tuple[str, List[str]]:
    if not path or not path.exists():
        return "", []
    raw = path.read_text(encoding="utf-8", errors="ignore")
    parser = TextExtractor()
    parser.feed(raw)
    text = parser.text()

    # Extract project-control/report terms that occur in conversations.
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9_%-]{2,}", text.lower())
    stop = {
        "the", "and", "for", "with", "that", "this", "from", "you", "your", "are", "was",
        "have", "has", "not", "can", "will", "but", "all", "only", "into", "make", "need",
        "create", "file", "report", "data", "same", "different", "inside", "also", "then",
    }
    counts = Counter(t for t in tokens if t not in stop)
    terms = [w for w, _ in counts.most_common(80)]
    return text, terms


class XlsxXmlReader:
    """Reads XLSX content directly from OOXML, including mildly malformed workbooks.

    This avoids depending on Excel's workbook metadata being perfect. It is intentionally
    read-only and never modifies the source workbook.
    """

    def __init__(self, path: Path) -> None:
        self.path = path
        self.shared_strings: List[str] = []
        self.sheet_map: List[Tuple[str, str]] = []
        self.chart_counts: Dict[str, int] = defaultdict(int)

    def _read_xml(self, z: zipfile.ZipFile, name: str) -> Optional[ET.Element]:
        try:
            return ET.fromstring(z.read(name))
        except KeyError:
            return None
        except ET.ParseError:
            return None

    def _load_shared_strings(self, z: zipfile.ZipFile) -> None:
        root = self._read_xml(z, "xl/sharedStrings.xml")
        if root is None:
            return
        out = []
        for si in root.findall("m:si", NS_MAIN):
            texts = [t.text or "" for t in si.findall(".//m:t", NS_MAIN)]
            out.append("".join(texts))
        self.shared_strings = out

    def _load_sheet_map(self, z: zipfile.ZipFile) -> None:
        wb = self._read_xml(z, "xl/workbook.xml")
        rels = self._read_xml(z, "xl/_rels/workbook.xml.rels")
        if wb is None:
            return
        rel_map: Dict[str, str] = {}
        if rels is not None:
            for rel in rels.findall("p:Relationship", NS_PKG_REL):
                rid = rel.attrib.get("Id", "")
                target = rel.attrib.get("Target", "")
                if rid and target:
                    if target.startswith("/"):
                        target = target.lstrip("/")
                    elif not target.startswith("xl/"):
                        target = "xl/" + target.lstrip("/")
                    rel_map[rid] = target

        sheets = wb.find("m:sheets", NS_MAIN)
        if sheets is None:
            return
        for idx, sh in enumerate(list(sheets), start=1):
            name = sh.attrib.get("name", f"Sheet{idx}")
            rid = sh.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
            target = rel_map.get(rid, f"xl/worksheets/sheet{idx}.xml")
            self.sheet_map.append((name, target))

    def _sheet_chart_count(self, z: zipfile.ZipFile, sheet_xml_path: str) -> int:
        # Find worksheet relationship to drawing, then count chart relationships.
        base = Path(sheet_xml_path)
        rel_path = str(base.parent / "_rels" / (base.name + ".rels")).replace("\\", "/")
        rels = self._read_xml(z, rel_path)
        if rels is None:
            return 0
        count = 0
        for rel in rels.findall("p:Relationship", NS_PKG_REL):
            rel_type = rel.attrib.get("Type", "")
            target = rel.attrib.get("Target", "")
            if rel_type.endswith("/drawing"):
                drawing_path = str((base.parent / target).resolve()).replace("\\", "/")
                # pathlib resolve is unsuitable for zip virtual paths; normalize manually.
                drawing_path = re.sub(r"^.*?/xl/", "xl/", drawing_path)
                if not drawing_path.startswith("xl/"):
                    # Common path ../drawings/drawing1.xml
                    drawing_path = "xl/" + str(Path("worksheets") / target).replace("\\", "/")
                    parts = []
                    for p in drawing_path.split("/"):
                        if p == "..":
                            if parts:
                                parts.pop()
                        elif p != ".":
                            parts.append(p)
                    drawing_path = "/".join(parts)
                dr_rels_path = str(Path(drawing_path).parent / "_rels" / (Path(drawing_path).name + ".rels")).replace("\\", "/")
                dr_rels = self._read_xml(z, dr_rels_path)
                if dr_rels is not None:
                    count += sum(1 for r in dr_rels.findall("p:Relationship", NS_PKG_REL)
                                 if r.attrib.get("Type", "").endswith("/chart"))
        return count

    def read(self, max_cells_per_sheet: int = 120000) -> Tuple[List[CellRecord], List[SheetProfile]]:
        if not zipfile.is_zipfile(self.path):
            raise ValueError(f"Not a valid .xlsx/.xlsm ZIP package: {self.path}")

        all_cells: List[CellRecord] = []
        profiles: List[SheetProfile] = []

        with zipfile.ZipFile(self.path, "r") as z:
            self._load_shared_strings(z)
            self._load_sheet_map(z)

            for sheet_name, xml_path in self.sheet_map:
                root = self._read_xml(z, xml_path)
                if root is None:
                    profiles.append(SheetProfile(name=sheet_name))
                    continue

                cells: List[CellRecord] = []
                row_max = col_max = 0
                numeric = formulas = texts = 0
                text_samples: List[str] = []

                for i, c in enumerate(root.findall(".//m:c", NS_MAIN)):
                    if i >= max_cells_per_sheet:
                        break
                    ref = c.attrib.get("r", "")
                    dtype = c.attrib.get("t")
                    f_node = c.find("m:f", NS_MAIN)
                    v_node = c.find("m:v", NS_MAIN)
                    formula = f_node.text if f_node is not None else None
                    value: Any = None

                    if dtype == "inlineStr":
                        t_node = c.find(".//m:t", NS_MAIN)
                        value = t_node.text if t_node is not None else ""
                    elif v_node is not None:
                        raw = v_node.text or ""
                        if dtype == "s":
                            try:
                                value = self.shared_strings[int(raw)]
                            except Exception:
                                value = raw
                        elif dtype in {"str", "e", "b"}:
                            value = raw
                        else:
                            try:
                                value = float(raw)
                                if value.is_integer():
                                    value = int(value)
                            except Exception:
                                value = raw

                    r = row_index(ref)
                    col = column_index(ref)
                    row_max = max(row_max, r)
                    col_max = max(col_max, col)
                    if formula:
                        formulas += 1
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        numeric += 1
                    elif value not in (None, ""):
                        texts += 1
                        if len(text_samples) < 300:
                            text_samples.append(str(value))

                    cells.append(CellRecord(sheet=sheet_name, ref=ref, row=r, col=col,
                                            value=value, formula=formula, data_type=dtype))

                total = len(cells)
                profile = SheetProfile(
                    name=sheet_name,
                    row_count=row_max,
                    col_count=col_max,
                    nonempty_cells=total,
                    numeric_cells=numeric,
                    formula_cells=formulas,
                    text_cells=texts,
                    chart_count=self._sheet_chart_count(z, xml_path),
                    numeric_density=(numeric / total if total else 0.0),
                    formula_density=(formulas / total if total else 0.0),
                    text_sample=" | ".join(text_samples)[:10000],
                )
                profiles.append(profile)
                all_cells.extend(cells)

        return all_cells, profiles


def best_metric_match(label: str, threshold: float = 0.58) -> Optional[Tuple[str, float]]:
    n = normalize_text(label)
    if not n or len(n) < 2:
        return None
    best: Optional[Tuple[str, float]] = None
    for key, aliases in METRIC_LEXICON.items():
        for alias in aliases:
            a = normalize_text(alias)
            if n == a:
                score = 1.0
            elif a in n or n in a:
                score = min(0.97, 0.72 + min(len(a), len(n)) / max(len(a), len(n)) * 0.22)
            else:
                score = difflib.SequenceMatcher(None, n, a).ratio()
            if best is None or score > best[1]:
                best = (key, score)
    return best if best and best[1] >= threshold else None


def discover_metrics(cells: Sequence[CellRecord]) -> Dict[str, Dict[str, Any]]:
    by_sheet_row: Dict[Tuple[str, int], List[CellRecord]] = defaultdict(list)
    for c in cells:
        by_sheet_row[(c.sheet, c.row)].append(c)
    for row_cells in by_sheet_row.values():
        row_cells.sort(key=lambda x: x.col)

    candidates: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for c in cells:
        if not isinstance(c.value, str) or not c.value.strip():
            continue
        match = best_metric_match(c.value)
        if not match:
            continue
        metric, score = match
        row_cells = by_sheet_row.get((c.sheet, c.row), [])
        nums = [x for x in row_cells if x.col > c.col and isinstance(x.value, (int, float)) and not isinstance(x.value, bool)]
        nums.sort(key=lambda x: x.col - c.col)
        nearest = nums[0] if nums else None
        candidates[metric].append({
            "sheet": c.sheet,
            "label_ref": c.ref,
            "label": c.value,
            "match_score": round(score, 4),
            "value_ref": nearest.ref if nearest else None,
            "value": nearest.value if nearest else None,
        })

    resolved: Dict[str, Dict[str, Any]] = {}
    for metric, items in candidates.items():
        items.sort(key=lambda x: (x["value"] is not None, x["match_score"]), reverse=True)
        resolved[metric] = {"best": items[0], "candidates": items[:20]}
    return resolved


def formula_signature(formula: str) -> str:
    f = formula.upper()
    f = re.sub(r"\$?[A-Z]{1,3}\$?\d+", "CELL", f)
    f = re.sub(r"\d+(?:\.\d+)?", "NUM", f)
    f = re.sub(r"\s+", "", f)
    return f[:180]


def summarize_formulas(cells: Sequence[CellRecord]) -> Dict[str, int]:
    counts = Counter()
    for c in cells:
        if c.formula:
            counts[formula_signature(c.formula)] += 1
    return dict(counts.most_common(100))


class MLDiscovery:
    def __init__(self, random_state: int = 42) -> None:
        self.random_state = random_state
        self.vectorizer = None
        self.clusterer = None
        self.anomaly = None
        self.feature_names: List[str] = []
        self.available = False

    def fit(self, profiles: List[SheetProfile], conversation_text: str = "") -> None:
        try:
            from sklearn.feature_extraction.text import TfidfVectorizer
            from sklearn.cluster import KMeans
            from sklearn.ensemble import IsolationForest
            import numpy as np
        except ImportError:
            self.available = False
            self._fallback_assign(profiles)
            return

        docs = []
        for p in profiles:
            docs.append((p.name + " " + p.text_sample)[:30000])
        if conversation_text:
            # Blend recurring conversation vocabulary into sheet discovery without letting
            # a large history overwhelm actual workbook text.
            convo_tail = conversation_text[-60000:]
            docs = [d + " " + convo_tail[:4000] for d in docs]

        if not docs:
            return

        self.vectorizer = TfidfVectorizer(
            lowercase=True,
            ngram_range=(1, 2),
            max_features=5000,
            sublinear_tf=True,
            token_pattern=r"(?u)\b[\w%.-]{2,}\b",
        )
        X = self.vectorizer.fit_transform(docs)
        n = len(profiles)
        k = max(1, min(8, int(round(math.sqrt(n))) if n > 1 else 1, n))
        self.clusterer = KMeans(n_clusters=k, random_state=self.random_state, n_init=10)
        labels = self.clusterer.fit_predict(X)

        numeric_features = np.array([
            [p.row_count, p.col_count, p.nonempty_cells, p.numeric_density,
             p.formula_density, p.chart_count]
            for p in profiles
        ], dtype=float)
        # stabilize scaling for structural features
        for j in range(numeric_features.shape[1]):
            col = numeric_features[:, j]
            sd = float(col.std())
            if sd > 1e-12:
                numeric_features[:, j] = (col - float(col.mean())) / sd

        if len(profiles) >= 5:
            self.anomaly = IsolationForest(contamination="auto", random_state=self.random_state)
            scores = self.anomaly.fit(numeric_features).decision_function(numeric_features)
        else:
            scores = [0.0] * len(profiles)

        self.feature_names = list(self.vectorizer.get_feature_names_out())
        for p, label, score in zip(profiles, labels, scores):
            p.cluster = int(label)
            p.anomaly_score = round(float(score), 6)
        self.available = True

    def _fallback_assign(self, profiles: List[SheetProfile]) -> None:
        for p in profiles:
            text = normalize_text(p.name + " " + p.text_sample)
            if any(x in text for x in ["dashboard", "summary", "overview"]):
                p.cluster = 0
            elif any(x in text for x in ["cash", "revenue", "profit"]):
                p.cluster = 1
            elif any(x in text for x in ["cost", "budget", "actual"]):
                p.cluster = 2
            else:
                p.cluster = 3
            p.anomaly_score = 0.0

    def save(self, out_dir: Path, profile: WorkbookProfile) -> None:
        out_dir.mkdir(parents=True, exist_ok=True)
        meta = {
            "app": APP_NAME,
            "version": VERSION,
            "sklearn_active": self.available,
            "workbook_profile": asdict(profile),
        }
        (out_dir / MODEL_META_FILENAME).write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")
        if self.available:
            try:
                import joblib
                joblib.dump({
                    "vectorizer": self.vectorizer,
                    "clusterer": self.clusterer,
                    "anomaly": self.anomaly,
                    "feature_names": self.feature_names,
                }, out_dir / MODEL_FILENAME)
            except Exception as exc:
                print(f"[WARN] Could not persist sklearn objects: {exc}")


class ConstrainedScenarioGenerator:
    """Generates coherent synthetic project-control values.

    The latent factors are intentionally correlated, then all derived KPIs are
    recalculated. This is more defensible than changing every metric independently.
    """

    def __init__(self, seed: int = 42) -> None:
        self.rng = random.Random(seed)

    def _positive(self, x: float, floor: float = 1.0) -> float:
        return max(floor, x)

    def generate(self, base: Dict[str, float], similarity: int) -> Dict[str, Any]:
        if similarity not in SIMILARITY_PROFILES:
            raise ValueError(f"similarity must be one of {sorted(SIMILARITY_PROFILES)}")

        # Lower visual similarity may use more different scenario figures while retaining
        # the same business meaning and relationships.
        magnitude = {90: 0.06, 60: 0.12, 30: 0.20, 10: 0.30}[similarity]

        contract = self._positive(base.get("contract_price", 100_000_000.0) * (1 + self.rng.uniform(-magnitude, magnitude)))
        direct_ratio = min(0.90, max(0.55, base.get("direct_budget", contract * 0.72) / max(contract, 1)))
        indirect_ratio = min(0.25, max(0.05, base.get("indirect_budget", contract * 0.12) / max(contract, 1)))
        fee_ratio = min(0.12, max(0.00, base.get("fees_budget", contract * 0.03) / max(contract, 1)))

        direct_budget = contract * direct_ratio * (1 + self.rng.uniform(-magnitude / 2, magnitude / 2))
        indirect_budget = contract * indirect_ratio * (1 + self.rng.uniform(-magnitude / 2, magnitude / 2))
        fees_budget = contract * fee_ratio * (1 + self.rng.uniform(-magnitude / 2, magnitude / 2))
        total_budget = direct_budget + indirect_budget + fees_budget

        base_progress = base.get("completion", 0.55)
        if base_progress > 1.5:
            base_progress /= 100.0
        progress = min(0.96, max(0.08, base_progress + self.rng.uniform(-0.12, 0.12)))

        # Earned value tied to execution budget and progress.
        earned_value = (direct_budget + indirect_budget) * progress

        # Cost pressure is the key latent factor driving CPI.
        target_cpi = min(1.18, max(0.72, base.get("cpi", 0.96) + self.rng.uniform(-magnitude, magnitude)))
        actual_cost = earned_value / target_cpi
        cost_variance = earned_value - actual_cost
        cpi = earned_value / actual_cost if actual_cost else 0.0

        billing_efficiency = min(1.15, max(0.72, 0.93 + self.rng.uniform(-magnitude, magnitude)))
        revenue = contract * progress * billing_efficiency
        # Include commercial burden already represented by fees_budget only proportionally to progress.
        recognized_fees = fees_budget * progress
        gross_profit = revenue - actual_cost - recognized_fees
        gross_profit_pct = gross_profit / revenue if revenue else 0.0

        # Cost mix sums exactly to actual cost.
        labels = ["Subcontractors", "Labor", "Equipment", "Materials", "Other"]
        raw = [self.rng.uniform(0.10, 0.35), self.rng.uniform(0.10, 0.28), self.rng.uniform(0.05, 0.18),
               self.rng.uniform(0.15, 0.40), self.rng.uniform(0.02, 0.12)]
        s = sum(raw)
        cost_mix = {k: actual_cost * v / s for k, v in zip(labels, raw)}

        # Indirect cost split sums to indirect actual cost estimate.
        indirect_actual = actual_cost * indirect_budget / max(direct_budget + indirect_budget, 1)
        soil_share = self.rng.uniform(0.10, 0.45)
        indirect_mix = {
            "Soil Settlement": indirect_actual * soil_share,
            "Other Indirect": indirect_actual * (1 - soil_share),
        }

        # Activity-level roll-up sums to direct budget and reconciles to EV/AC at component level.
        activity_names = [
            "Mobilization & Preliminaries", "Earthworks", "Substructure", "Superstructure",
            "Architectural Works", "MEP Works", "External Works", "Testing & Handover"
        ]
        weights_raw = [self.rng.uniform(0.7, 1.4) for _ in activity_names]
        ws = sum(weights_raw)
        weights = [w / ws for w in weights_raw]
        activities = []
        remaining_ev_target = earned_value * direct_budget / max(direct_budget + indirect_budget, 1)
        direct_actual_target = actual_cost * direct_budget / max(direct_budget + indirect_budget, 1)

        ev_alloc = []
        ac_alloc = []
        for i, (name, w) in enumerate(zip(activity_names, weights)):
            b = direct_budget * w
            local_progress = min(1.0, max(0.0, progress + self.rng.uniform(-0.20, 0.20)))
            e = b * local_progress
            local_cpi = min(1.25, max(0.65, cpi + self.rng.uniform(-0.16, 0.16)))
            a = e / local_cpi if local_cpi else e
            ev_alloc.append(e)
            ac_alloc.append(a)
            activities.append({"activity": name, "weight": w, "budget": b, "completion": local_progress, "ev": e, "ac": a})

        # Scale activity EV and AC to exact component targets while preserving relative signals.
        ev_scale = remaining_ev_target / max(sum(ev_alloc), 1)
        ac_scale = direct_actual_target / max(sum(ac_alloc), 1)
        for row in activities:
            row["ev"] *= ev_scale
            row["ac"] *= ac_scale
            row["cv"] = row["ev"] - row["ac"]
            row["cpi"] = row["ev"] / row["ac"] if row["ac"] else 0.0
            row["status"] = "Favorable" if row["cpi"] >= 1.0 else "Adverse"

        # Quarterly cash profile sums to revenue and actual cost.
        quarters = ["Q1", "Q2", "Q3", "Q4"]
        in_raw = [self.rng.uniform(0.10, 0.30), self.rng.uniform(0.18, 0.34), self.rng.uniform(0.22, 0.38), self.rng.uniform(0.12, 0.30)]
        out_raw = [self.rng.uniform(0.12, 0.30), self.rng.uniform(0.20, 0.34), self.rng.uniform(0.22, 0.38), self.rng.uniform(0.12, 0.28)]
        in_sum, out_sum = sum(in_raw), sum(out_raw)
        cashflow = []
        cum_in = cum_out = 0.0
        for q, ri, ro in zip(quarters, in_raw, out_raw):
            ci = revenue * ri / in_sum
            co = actual_cost * ro / out_sum
            cum_in += ci
            cum_out += co
            cashflow.append({"period": q, "cash_in": ci, "cash_out": co,
                             "net": ci - co, "cumulative_net": cum_in - cum_out})

        return {
            "similarity": similarity,
            "profile": SIMILARITY_PROFILES[similarity],
            "contract_price": contract,
            "direct_budget": direct_budget,
            "indirect_budget": indirect_budget,
            "fees_budget": fees_budget,
            "total_budget": total_budget,
            "completion": progress,
            "earned_value": earned_value,
            "actual_cost": actual_cost,
            "cost_variance": cost_variance,
            "cpi": cpi,
            "status": "Favorable" if cpi >= 1.0 else "Adverse",
            "revenue": revenue,
            "gross_profit": gross_profit,
            "gross_profit_pct": gross_profit_pct,
            "cost_mix": cost_mix,
            "indirect_mix": indirect_mix,
            "activities": activities,
            "cashflow": cashflow,
        }


def extract_base_numeric(metrics: Dict[str, Dict[str, Any]]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    mapping = {
        "contract_price": "contract_price", "direct_budget": "direct_budget",
        "indirect_budget": "indirect_budget", "fees_budget": "fees_budget",
        "total_budget": "total_budget", "earned_value": "earned_value",
        "actual_cost": "actual_cost", "cost_variance": "cost_variance",
        "cpi": "cpi", "completion": "completion", "revenue": "revenue",
        "gross_profit": "gross_profit", "gross_profit_pct": "gross_profit_pct",
    }
    for src, dst in mapping.items():
        item = metrics.get(src, {}).get("best", {})
        value = item.get("value")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            out[dst] = float(value)
    return out


def validate_scenario(s: Dict[str, Any], tolerance: float = 1e-6) -> List[str]:
    errors: List[str] = []
    def close(a: float, b: float, rel: float = 1e-6) -> bool:
        return abs(a - b) <= max(tolerance, rel * max(abs(a), abs(b), 1.0))

    if not close(s["total_budget"], s["direct_budget"] + s["indirect_budget"] + s["fees_budget"]):
        errors.append("Total budget does not reconcile to direct + indirect + fees.")
    if not close(s["cost_variance"], s["earned_value"] - s["actual_cost"]):
        errors.append("CV does not equal EV - AC.")
    if s["actual_cost"] and not close(s["cpi"], s["earned_value"] / s["actual_cost"]):
        errors.append("CPI does not equal EV / AC.")
    if s["revenue"] and not close(s["gross_profit_pct"], s["gross_profit"] / s["revenue"]):
        errors.append("GP% does not equal GP / revenue.")
    if not close(sum(s["cost_mix"].values()), s["actual_cost"]):
        errors.append("Actual cost mix does not sum to actual cost.")
    if not close(sum(x["cash_in"] for x in s["cashflow"]), s["revenue"]):
        errors.append("Cash-in does not sum to revenue.")
    if not close(sum(x["cash_out"] for x in s["cashflow"]), s["actual_cost"]):
        errors.append("Cash-out does not sum to actual cost.")
    if not close(sum(x["budget"] for x in s["activities"]), s["direct_budget"]):
        errors.append("Activity budgets do not sum to direct budget.")
    return errors


def format_money(x: float) -> str:
    a = abs(x)
    if a >= 1_000_000_000:
        return f"{x/1_000_000_000:,.2f}B"
    if a >= 1_000_000:
        return f"{x/1_000_000:,.2f}M"
    if a >= 1_000:
        return f"{x/1_000:,.2f}K"
    return f"{x:,.0f}"


def generate_xlsx(s: Dict[str, Any], out_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("XLSX generation requires openpyxl. Install with: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    data = wb.create_sheet("Scenario Data")
    acts = wb.create_sheet("Activity Performance")
    cash = wb.create_sheet("Cash Flow")

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    signal_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"Synthetic Cost Control Report — {s['similarity']}% Similarity"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:H1")
    ws["A2"] = f"Design family: {s['profile']['name']} | Generated by {APP_NAME} v{VERSION}"
    ws.merge_cells("A2:H2")

    kpis = [
        ("Contract Price", s["contract_price"]), ("Total Budget", s["total_budget"]),
        ("Earned Value", s["earned_value"]), ("Actual Cost", s["actual_cost"]),
        ("Cost Variance", s["cost_variance"]), ("CPI", s["cpi"]),
        ("Revenue", s["revenue"]), ("Gross Profit %", s["gross_profit_pct"]),
    ]

    # Visual divergence is achieved mainly by arrangement, not by changing metric semantics.
    layout = s["profile"]["layout"]
    if layout == "matrix":
        positions = [(4,1),(4,3),(4,5),(4,7),(7,1),(7,3),(7,5),(7,7)]
    elif layout == "executive":
        positions = [(4,1),(4,3),(4,5),(4,7),(6,1),(6,3),(6,5),(6,7)]
    elif layout == "control_room":
        positions = [(4,1),(6,1),(8,1),(10,1),(4,3),(6,3),(8,3),(10,3)]
    else:
        positions = [(4,1),(4,4),(7,1),(7,4),(10,1),(10,4),(13,1),(13,4)]

    for (label, value), (r, c) in zip(kpis, positions):
        ws.cell(r, c, label)
        ws.cell(r, c).font = Font(bold=True)
        ws.cell(r, c).fill = header_fill
        ws.cell(r, c).border = border
        ws.cell(r+1, c, value)
        ws.cell(r+1, c).border = border
        if label in {"CPI"}:
            ws.cell(r+1, c).number_format = "0.00"
        elif "%" in label:
            ws.cell(r+1, c).number_format = "0.0%"
        else:
            ws.cell(r+1, c).number_format = "#,##0"
        if c < 8:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=min(c+1,8))
            ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=min(c+1,8))

    # Scenario data block.
    data.append(["Metric", "Value"])
    scalar_keys = ["contract_price","direct_budget","indirect_budget","fees_budget","total_budget",
                   "completion","earned_value","actual_cost","cost_variance","cpi","revenue","gross_profit","gross_profit_pct"]
    for key in scalar_keys:
        data.append([key, s[key]])
    start_mix = data.max_row + 2
    data.cell(start_mix, 1, "Cost Type")
    data.cell(start_mix, 2, "Actual Cost")
    for i, (k, v) in enumerate(s["cost_mix"].items(), start=start_mix+1):
        data.cell(i, 1, k); data.cell(i, 2, v)

    # Activities.
    acts.append(["Activity", "Weight", "Budget", "Completion", "EV", "AC", "CV", "CPI", "Status"])
    for row in s["activities"]:
        acts.append([row["activity"], row["weight"], row["budget"], row["completion"], row["ev"], row["ac"], row["cv"], row["cpi"], row["status"]])
    acts.freeze_panes = "A2"

    # Cash flow.
    cash.append(["Period", "Cash In", "Cash Out", "Net", "Cumulative Net"])
    for row in s["cashflow"]:
        cash.append([row["period"], row["cash_in"], row["cash_out"], row["net"], row["cumulative_net"]])

    # Styling.
    for sh in [data, acts, cash]:
        for cell in sh[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
        for row in sh.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for col in range(1, sh.max_column + 1):
            max_len = max(len(str(sh.cell(r, col).value or "")) for r in range(1, sh.max_row + 1))
            sh.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 34)

    for row in acts.iter_rows(min_row=2, min_col=2, max_col=8):
        for cell in row:
            if cell.column in {2,4}:
                cell.number_format = "0.0%"
            elif cell.column == 8:
                cell.number_format = "0.00"
            else:
                cell.number_format = "#,##0"
    for row in cash.iter_rows(min_row=2, min_col=2, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"

    # Charts. Each profile uses a different arrangement/type emphasis.
    chart_anchor_map = {
        "matrix": ("A11", "E11", "A27"),
        "executive": ("A10", "E10", "A26"),
        "control_room": ("E4", "E19", "A26"),
        "decision_lens": ("F4", "F20", "A29"),
    }
    a1, a2, a3 = chart_anchor_map[layout]

    mix_chart = DoughnutChart() if s["similarity"] >= 60 else BarChart()
    mix_chart.title = "Actual Cost Mix"
    mix_data = Reference(data, min_col=2, min_row=start_mix, max_row=start_mix+len(s["cost_mix"]))
    mix_cats = Reference(data, min_col=1, min_row=start_mix+1, max_row=start_mix+len(s["cost_mix"]))
    mix_chart.add_data(mix_data, titles_from_data=True)
    mix_chart.set_categories(mix_cats)
    mix_chart.height = 7.2; mix_chart.width = 10.5
    ws.add_chart(mix_chart, a1)

    perf_chart = BarChart()
    perf_chart.type = "bar" if s["similarity"] <= 60 else "col"
    perf_chart.title = "Activity EV vs AC"
    perf_data = Reference(acts, min_col=5, max_col=6, min_row=1, max_row=acts.max_row)
    perf_cats = Reference(acts, min_col=1, min_row=2, max_row=acts.max_row)
    perf_chart.add_data(perf_data, titles_from_data=True)
    perf_chart.set_categories(perf_cats)
    perf_chart.height = 8; perf_chart.width = 11
    ws.add_chart(perf_chart, a2)

    cf_chart = LineChart()
    cf_chart.title = "Cash In vs Cash Out"
    cf_data = Reference(cash, min_col=2, max_col=3, min_row=1, max_row=cash.max_row)
    cf_cats = Reference(cash, min_col=1, min_row=2, max_row=cash.max_row)
    cf_chart.add_data(cf_data, titles_from_data=True)
    cf_chart.set_categories(cf_cats)
    cf_chart.height = 8; cf_chart.width = 12
    ws.add_chart(cf_chart, a3)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A3"

    # Add reconciliation notes, making the synthetic nature explicit.
    note_row = max(ws.max_row + 2, 43)
    ws.cell(note_row, 1, "Control checks")
    ws.cell(note_row, 1).font = Font(bold=True)
    checks = [
        f"Total budget = Direct + Indirect + Fees: {format_money(s['total_budget'])}",
        f"CV = EV - AC: {format_money(s['cost_variance'])}",
        f"CPI = EV / AC: {s['cpi']:.3f}",
        f"Cost mix reconciles to AC: {format_money(sum(s['cost_mix'].values()))}",
        f"Cash In reconciles to Revenue: {format_money(sum(x['cash_in'] for x in s['cashflow']))}",
        f"Cash Out reconciles to AC: {format_money(sum(x['cash_out'] for x in s['cashflow']))}",
    ]
    for i, t in enumerate(checks, start=note_row+1):
        ws.cell(i, 1, t)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def create_html_audit(profile: WorkbookProfile, scenarios: List[Dict[str, Any]], out_path: Path) -> None:
    context = {
        "app": APP_NAME, "version": VERSION, "profile": asdict(profile),
        "scenarios": scenarios, "generated_at": datetime.now().isoformat(timespec="seconds")
    }
    template = """<!doctype html>
<html><head><meta charset='utf-8'><title>{{ app }} Audit</title>
<style>body{font-family:Segoe UI,Arial,sans-serif;margin:32px;background:#f6f8fb;color:#1d2733}.card{background:#fff;border:1px solid #dde3ea;border-radius:12px;padding:18px;margin:12px 0}table{border-collapse:collapse;width:100%}th,td{border-bottom:1px solid #e7ebef;padding:8px;text-align:left}th{background:#eef4f9}.ok{font-weight:700}</style>
</head><body><h1>{{ app }}</h1><p>Version {{ version }} — {{ generated_at }}</p>
<div class='card'><h2>Workbook</h2><p>{{ profile.path }}</p><p>SHA256: {{ profile.sha256 }}</p><p>Sheets: {{ profile.sheet_names|length }}</p></div>
<div class='card'><h2>Discovered Metrics</h2><table><tr><th>Metric</th><th>Sheet</th><th>Label</th><th>Value</th></tr>
{% for k,v in profile.metrics.items() %}<tr><td>{{ k }}</td><td>{{ v.best.sheet }}</td><td>{{ v.best.label }}</td><td>{{ v.best.value }}</td></tr>{% endfor %}</table></div>
<div class='card'><h2>Synthetic Scenarios</h2><table><tr><th>Similarity</th><th>CPI</th><th>CV</th><th>Revenue</th><th>GP%</th><th>Status</th></tr>
{% for s in scenarios %}<tr><td>{{ s.similarity }}%</td><td>{{ '%.3f'|format(s.cpi) }}</td><td>{{ '%.0f'|format(s.cost_variance) }}</td><td>{{ '%.0f'|format(s.revenue) }}</td><td>{{ '%.1f%%'|format(s.gross_profit_pct*100) }}</td><td class='ok'>{{ s.status }}</td></tr>{% endfor %}</table></div>
</body></html>"""
    try:
        from jinja2 import Template
        rendered = Template(template).render(**context)
    except ImportError:
        # Minimal dependency-free fallback.
        rows = "".join(
            f"<tr><td>{s['similarity']}%</td><td>{s['cpi']:.3f}</td><td>{s['cost_variance']:.0f}</td><td>{s['revenue']:.0f}</td><td>{s['gross_profit_pct']:.1%}</td><td>{s['status']}</td></tr>"
            for s in scenarios
        )
        rendered = f"<!doctype html><meta charset='utf-8'><title>{APP_NAME}</title><h1>{APP_NAME}</h1><p>{html.escape(profile.path)}</p><table><tr><th>Similarity</th><th>CPI</th><th>CV</th><th>Revenue</th><th>GP%</th><th>Status</th></tr>{rows}</table>"
    out_path.write_text(rendered, encoding="utf-8")


def analyze_workbook(xlsx_path: Path, html_path: Optional[Path] = None) -> Tuple[WorkbookProfile, MLDiscovery, str]:
    conversation_text, terms = load_conversation_html(html_path)
    reader = XlsxXmlReader(xlsx_path)
    cells, profiles = reader.read()
    metrics = discover_metrics(cells)
    formulas = summarize_formulas(cells)

    # Annotate each sheet with metric names found there.
    metric_by_sheet: Dict[str, set] = defaultdict(set)
    for metric, info in metrics.items():
        for c in info.get("candidates", []):
            metric_by_sheet[c["sheet"]].add(metric)
    for p in profiles:
        p.discovered_metrics = sorted(metric_by_sheet.get(p.name, set()))

    ml = MLDiscovery(random_state=42)
    ml.fit(profiles, conversation_text)

    profile = WorkbookProfile(
        path=str(xlsx_path.resolve()),
        sha256=sha256_file(xlsx_path),
        sheet_names=[p.name for p in profiles],
        sheets=profiles,
        metrics=metrics,
        formulas=formulas,
        conversation_terms=terms,
    )
    return profile, ml, conversation_text


def write_analysis(profile: WorkbookProfile, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(asdict(profile), indent=2, ensure_ascii=False), encoding="utf-8")


def run_analyze(args: argparse.Namespace) -> int:
    xlsx = Path(args.xlsx)
    html_path = Path(args.html) if args.html else (xlsx.parent / DEFAULT_HTML)
    if not html_path.exists():
        html_path = None
    profile, ml, _ = analyze_workbook(xlsx, html_path)
    out = Path(args.out or "report_analysis.json")
    write_analysis(profile, out)
    print(f"Analysis written: {out.resolve()}")
    print(f"Sheets: {len(profile.sheet_names)} | Metrics discovered: {len(profile.metrics)} | sklearn active: {ml.available}")
    if html_path:
        print(f"Conversation source: {html_path.resolve()} | recurring terms learned: {len(profile.conversation_terms)}")
    else:
        print(f"Conversation source not found; place {DEFAULT_HTML} beside the script/workbook to include it.")
    return 0


def run_learn(args: argparse.Namespace) -> int:
    xlsx = Path(args.xlsx)
    html_path = Path(args.html) if args.html else (xlsx.parent / DEFAULT_HTML)
    if not html_path.exists():
        html_path = None
    profile, ml, _ = analyze_workbook(xlsx, html_path)
    out_dir = Path(args.out or "report_ml_model")
    ml.save(out_dir, profile)
    print(f"Model package written: {out_dir.resolve()}")
    return 0


def run_generate(args: argparse.Namespace) -> int:
    xlsx = Path(args.xlsx)
    html_path = Path(args.html) if args.html else (xlsx.parent / DEFAULT_HTML)
    if not html_path.exists():
        html_path = None
    profile, ml, _ = analyze_workbook(xlsx, html_path)
    out_dir = Path(args.out or "generated_reports")
    out_dir.mkdir(parents=True, exist_ok=True)

    base = extract_base_numeric(profile.metrics)
    generator = ConstrainedScenarioGenerator(seed=args.seed)
    scenarios: List[Dict[str, Any]] = []
    for similarity in [90, 60, 30, 10]:
        scenario = generator.generate(base, similarity)
        errors = validate_scenario(scenario)
        if errors:
            raise RuntimeError(f"Scenario {similarity}% failed validation: {'; '.join(errors)}")
        scenarios.append(scenario)
        json_path = out_dir / f"scenario_{similarity}.json"
        json_path.write_text(json.dumps(scenario, indent=2, ensure_ascii=False), encoding="utf-8")
        xlsx_out = out_dir / f"synthetic_report_{similarity}_similarity.xlsx"
        generate_xlsx(scenario, xlsx_out)
        print(f"Created {similarity}%: {xlsx_out.resolve()}")

    write_analysis(profile, out_dir / "source_report_analysis.json")
    ml.save(out_dir / "model", profile)
    create_html_audit(profile, scenarios, out_dir / "ml_audit.html")
    print(f"Audit: {(out_dir / 'ml_audit.html').resolve()}")
    return 0


def interactive() -> int:
    print(f"\n{APP_NAME} v{VERSION}")
    print("=" * 64)
    print("This engine reads the source XLSX without changing it.")
    cwd = Path.cwd()
    default_html = cwd / DEFAULT_HTML

    xlsx_candidates = sorted(list(cwd.glob("*.xlsx")) + list(cwd.glob("*.xlsm")))
    if not xlsx_candidates:
        print("No .xlsx/.xlsm found in the current folder.")
        raw = input("Enter full Excel path: ").strip().strip('"')
        xlsx = Path(raw)
    else:
        print("Excel files:")
        for i, p in enumerate(xlsx_candidates, 1):
            print(f"  {i}. {p.name}")
        raw = input("Select file number [1]: ").strip()
        idx = int(raw or "1") - 1
        xlsx = xlsx_candidates[idx]

    if not xlsx.exists():
        print(f"File not found: {xlsx}")
        return 2

    html_path = default_html if default_html.exists() else None
    if html_path:
        print(f"Conversation learning source detected: {html_path.name}")
    else:
        print(f"Optional {DEFAULT_HTML} not found. Excel structural learning will still run.")

    print("\n1. Analyze only")
    print("2. Learn/save model")
    print("3. Generate 90/60/30/10 synthetic reports")
    choice = input("Choose [3]: ").strip() or "3"

    class A: pass
    a = A(); a.xlsx = str(xlsx); a.html = str(html_path) if html_path else None; a.seed = 42
    if choice == "1":
        a.out = "report_analysis.json"
        return run_analyze(a)
    if choice == "2":
        a.out = "report_ml_model"
        return run_learn(a)
    a.out = "generated_reports"
    return run_generate(a)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description=APP_NAME)
    sub = p.add_subparsers(dest="command")

    a = sub.add_parser("analyze", help="Analyze workbook structure, metrics and formulas")
    a.add_argument("xlsx")
    a.add_argument("--html", help=f"Optional conversation HTML; default: {DEFAULT_HTML} beside workbook")
    a.add_argument("--out", default="report_analysis.json")
    a.set_defaults(func=run_analyze)

    l = sub.add_parser("learn", help="Train unsupervised structural/semantic model")
    l.add_argument("xlsx")
    l.add_argument("--html", help=f"Optional conversation HTML; default: {DEFAULT_HTML} beside workbook")
    l.add_argument("--out", default="report_ml_model")
    l.set_defaults(func=run_learn)

    g = sub.add_parser("generate", help="Generate coherent 90/60/30/10 synthetic report variants")
    g.add_argument("xlsx")
    g.add_argument("--html", help=f"Optional conversation HTML; default: {DEFAULT_HTML} beside workbook")
    g.add_argument("--out", default="generated_reports")
    g.add_argument("--seed", type=int, default=42)
    g.set_defaults(func=run_generate)
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    if not argv:
        return interactive()
    parser = build_parser()
    args = parser.parse_args(argv)
    if not getattr(args, "command", None):
        parser.print_help()
        return 1
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
